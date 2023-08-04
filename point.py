import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def update_existing_sheet(existing_sheet_filename, input_sheet_filename):
    try:
        existing_workbook = load_workbook(existing_sheet_filename)
        existing_sheet = existing_workbook.active
        new_data = pd.read_excel(input_sheet_filename)

        name_points_dict = new_data.groupby('Name')['Points'].sum().to_dict()

        for name, points in name_points_dict.items():
            existing_row = existing_sheet.find(name)

            if not existing_row:
                new_row = [name, points]
                existing_sheet.append(new_row)
            else:
                row_num = existing_row.row
                existing_sheet.cell(row=row_num, column=2).value += points

        existing_workbook.save(existing_sheet_filename)
        print("Data successfully updated in the existing sheet.")
    except FileNotFoundError:
        print("Error: One or both of the input files not found.")
    except Exception as e:
        print(f"Error occurred while updating the sheet: {e}")

existing_sheet_filename = 'existing_sheet.xlsx'
input_sheet_filename = 'input_sheet.xlsx'
update_existing_sheet(existing_sheet_filename, input_sheet_filename)
